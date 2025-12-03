"""
Report Generator Module - PDF and Excel Export
===============================================
Comprehensive report generation for Solar PV Testing LIMS-QMS.

Features:
- PDF report generation with company branding
- Excel data export with multiple sheets
- Test result documentation
- Summary statistics
- Digital signature placeholders
- Support for multiple test formats
"""

import io
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
import streamlit as st
import pandas as pd

# PDF Generation Libraries
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, Image, HRFlowable
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# Alternative PDF library
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# Excel Generation
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class CompanySettings:
    """Company branding and settings for reports"""

    DEFAULT_SETTINGS = {
        'company_name': 'Solar PV Testing Laboratory',
        'company_address': '123 Solar Drive, Renewable City, RC 12345',
        'company_phone': '+1 (555) 123-4567',
        'company_email': 'info@solarpvtesting.com',
        'company_website': 'www.solarpvtesting.com',
        'logo_path': None,
        'accreditation': 'ISO/IEC 17025:2017 Accredited',
        'lab_code': 'SPVT-LAB-001'
    }

    @classmethod
    def get_settings(cls) -> Dict[str, Any]:
        """Get company settings from session state or defaults"""
        if 'company_settings' in st.session_state:
            return {**cls.DEFAULT_SETTINGS, **st.session_state.company_settings}
        return cls.DEFAULT_SETTINGS.copy()


class PDFReportGenerator:
    """Generate professional PDF reports using ReportLab"""

    def __init__(self):
        self.company = CompanySettings.get_settings()
        self.styles = None
        if REPORTLAB_AVAILABLE:
            self.styles = getSampleStyleSheet()
            self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom paragraph styles"""
        if not self.styles:
            return

        # Title style
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=12,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#FF6B35')
        ))

        # Subtitle style
        self.styles.add(ParagraphStyle(
            name='ReportSubtitle',
            parent=self.styles['Normal'],
            fontSize=12,
            spaceAfter=6,
            alignment=TA_CENTER,
            textColor=colors.grey
        ))

        # Section header
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceBefore=12,
            spaceAfter=6,
            textColor=colors.HexColor('#1f77b4')
        ))

        # Table header
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold'
        ))

    def generate_test_report(
        self,
        test_data: Dict[str, Any],
        results_data: List[Dict[str, Any]],
        include_charts: bool = True
    ) -> Optional[bytes]:
        """
        Generate a comprehensive PDF test report

        Args:
            test_data: Dictionary containing test metadata
            results_data: List of test results
            include_charts: Whether to include visual charts

        Returns:
            PDF bytes or None if generation fails
        """
        if not REPORTLAB_AVAILABLE:
            st.error("PDF generation requires reportlab. Install with: pip install reportlab")
            return None

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )

        elements = []

        # Header Section
        elements.extend(self._create_header())

        # Report Title
        title = test_data.get('title', 'Test Report')
        elements.append(Paragraph(title, self.styles['ReportTitle']))
        elements.append(Paragraph(
            f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            self.styles['ReportSubtitle']
        ))
        elements.append(Spacer(1, 0.25*inch))

        # Test Information Section
        elements.extend(self._create_test_info_section(test_data))

        # Results Summary Section
        elements.extend(self._create_results_summary(results_data))

        # Detailed Results Table
        elements.extend(self._create_results_table(results_data))

        # Pass/Fail Status Section
        elements.extend(self._create_status_section(test_data, results_data))

        # Digital Signature Section
        elements.extend(self._create_signature_section())

        # Footer
        elements.extend(self._create_footer())

        # Build PDF
        try:
            doc.build(elements)
            buffer.seek(0)
            return buffer.getvalue()
        except Exception as e:
            st.error(f"Error generating PDF: {e}")
            return None

    def _create_header(self) -> List:
        """Create report header with company branding"""
        elements = []

        # Company name and details
        header_data = [
            [self.company['company_name']],
            [self.company['company_address']],
            [f"Tel: {self.company['company_phone']} | Email: {self.company['company_email']}"],
            [self.company['accreditation']]
        ]

        header_table = Table(header_data, colWidths=[6*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (0, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (0, 0), 14),
            ('FONTSIZE', (0, 1), (0, -1), 9),
            ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#FF6B35')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))

        elements.append(header_table)
        elements.append(Spacer(1, 0.1*inch))
        elements.append(HRFlowable(
            width="100%",
            thickness=2,
            color=colors.HexColor('#FF6B35'),
            spaceBefore=0.1*inch,
            spaceAfter=0.2*inch
        ))

        return elements

    def _create_test_info_section(self, test_data: Dict[str, Any]) -> List:
        """Create test information section"""
        elements = []
        elements.append(Paragraph("Test Information", self.styles['SectionHeader']))

        info_data = [
            ['Field', 'Value'],
            ['Protocol ID', test_data.get('protocol_id', 'N/A')],
            ['Protocol Name', test_data.get('protocol_name', 'N/A')],
            ['Sample ID', test_data.get('sample_id', 'N/A')],
            ['Test Date', test_data.get('test_date', datetime.now().strftime('%Y-%m-%d'))],
            ['Operator', test_data.get('operator', 'N/A')],
            ['Equipment Used', test_data.get('equipment', 'N/A')],
            ['Environmental Conditions', test_data.get('conditions', 'Standard Laboratory Conditions')],
        ]

        table = Table(info_data, colWidths=[2*inch, 4*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')])
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.25*inch))
        return elements

    def _create_results_summary(self, results_data: List[Dict[str, Any]]) -> List:
        """Create results summary section"""
        elements = []
        elements.append(Paragraph("Results Summary", self.styles['SectionHeader']))

        if not results_data:
            elements.append(Paragraph("No results data available.", self.styles['Normal']))
            return elements

        # Calculate summary statistics
        total_tests = len(results_data)
        passed = sum(1 for r in results_data if r.get('status', '').upper() == 'PASS')
        failed = sum(1 for r in results_data if r.get('status', '').upper() == 'FAIL')
        pending = total_tests - passed - failed

        summary_data = [
            ['Metric', 'Value'],
            ['Total Measurements', str(total_tests)],
            ['Passed', str(passed)],
            ['Failed', str(failed)],
            ['Pending/Other', str(pending)],
            ['Pass Rate', f"{(passed/total_tests*100):.1f}%" if total_tests > 0 else 'N/A']
        ]

        table = Table(summary_data, colWidths=[2*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#28a745')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))

        elements.append(table)
        elements.append(Spacer(1, 0.25*inch))
        return elements

    def _create_results_table(self, results_data: List[Dict[str, Any]]) -> List:
        """Create detailed results table"""
        elements = []
        elements.append(Paragraph("Detailed Test Results", self.styles['SectionHeader']))

        if not results_data:
            elements.append(Paragraph("No detailed results available.", self.styles['Normal']))
            return elements

        # Build table data
        headers = ['#', 'Parameter', 'Value', 'Unit', 'Spec Min', 'Spec Max', 'Status']
        table_data = [headers]

        for i, result in enumerate(results_data, 1):
            status = result.get('status', 'N/A').upper()
            status_display = status

            row = [
                str(i),
                result.get('parameter', 'N/A'),
                str(result.get('value', 'N/A')),
                result.get('unit', ''),
                str(result.get('spec_min', '-')),
                str(result.get('spec_max', '-')),
                status_display
            ]
            table_data.append(row)

        table = Table(table_data, colWidths=[0.4*inch, 1.5*inch, 1*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch])

        # Style the table
        style_commands = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]

        # Color code status column
        for i, result in enumerate(results_data, 1):
            status = result.get('status', '').upper()
            if status == 'PASS':
                style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#d4edda')))
                style_commands.append(('TEXTCOLOR', (-1, i), (-1, i), colors.HexColor('#155724')))
            elif status == 'FAIL':
                style_commands.append(('BACKGROUND', (-1, i), (-1, i), colors.HexColor('#f8d7da')))
                style_commands.append(('TEXTCOLOR', (-1, i), (-1, i), colors.HexColor('#721c24')))

        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        elements.append(Spacer(1, 0.25*inch))
        return elements

    def _create_status_section(self, test_data: Dict[str, Any], results_data: List[Dict[str, Any]]) -> List:
        """Create overall pass/fail status section"""
        elements = []
        elements.append(Paragraph("Overall Test Status", self.styles['SectionHeader']))

        # Determine overall status
        if not results_data:
            overall_status = 'INCOMPLETE'
            status_color = colors.HexColor('#ffc107')
        else:
            failed = any(r.get('status', '').upper() == 'FAIL' for r in results_data)
            if failed:
                overall_status = 'FAIL'
                status_color = colors.HexColor('#dc3545')
            else:
                overall_status = 'PASS'
                status_color = colors.HexColor('#28a745')

        status_table = Table(
            [[f"OVERALL RESULT: {overall_status}"]],
            colWidths=[4*inch]
        )
        status_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), status_color),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.white),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 16),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))

        elements.append(status_table)
        elements.append(Spacer(1, 0.25*inch))

        # Comments section
        comments = test_data.get('comments', '')
        if comments:
            elements.append(Paragraph("Comments:", self.styles['Normal']))
            elements.append(Paragraph(comments, self.styles['Normal']))
            elements.append(Spacer(1, 0.25*inch))

        return elements

    def _create_signature_section(self) -> List:
        """Create digital signature placeholder section"""
        elements = []
        elements.append(Spacer(1, 0.5*inch))
        elements.append(Paragraph("Authorization", self.styles['SectionHeader']))

        sig_data = [
            ['Tested By:', '_' * 30, 'Date:', '_' * 15],
            ['Reviewed By:', '_' * 30, 'Date:', '_' * 15],
            ['Approved By:', '_' * 30, 'Date:', '_' * 15],
        ]

        sig_table = Table(sig_data, colWidths=[1*inch, 2*inch, 0.6*inch, 1.2*inch])
        sig_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('TOPPADDING', (0, 0), (-1, -1), 15),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ]))

        elements.append(sig_table)
        return elements

    def _create_footer(self) -> List:
        """Create report footer"""
        elements = []
        elements.append(Spacer(1, 0.5*inch))
        elements.append(HRFlowable(
            width="100%",
            thickness=1,
            color=colors.grey,
            spaceBefore=0.1*inch,
            spaceAfter=0.1*inch
        ))

        footer_text = f"""
        This report is generated by {self.company['company_name']} LIMS-QMS System.
        Lab Code: {self.company['lab_code']} | {self.company['accreditation']}
        Report ID: RPT-{datetime.now().strftime('%Y%m%d%H%M%S')}
        """

        elements.append(Paragraph(footer_text, ParagraphStyle(
            name='Footer',
            fontSize=8,
            textColor=colors.grey,
            alignment=TA_CENTER
        )))

        return elements


class FPDFReportGenerator:
    """Alternative PDF generator using FPDF2"""

    def __init__(self):
        self.company = CompanySettings.get_settings()

    def generate_test_report(
        self,
        test_data: Dict[str, Any],
        results_data: List[Dict[str, Any]]
    ) -> Optional[bytes]:
        """Generate PDF report using FPDF"""
        if not FPDF_AVAILABLE:
            st.error("PDF generation requires fpdf2. Install with: pip install fpdf2")
            return None

        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Header
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_text_color(255, 107, 53)  # Orange
        pdf.cell(0, 10, self.company['company_name'], ln=True, align='C')

        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, self.company['company_address'], ln=True, align='C')
        pdf.cell(0, 6, f"Tel: {self.company['company_phone']} | {self.company['company_email']}", ln=True, align='C')
        pdf.cell(0, 6, self.company['accreditation'], ln=True, align='C')

        pdf.ln(5)
        pdf.set_draw_color(255, 107, 53)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(10)

        # Title
        pdf.set_font('Helvetica', 'B', 14)
        pdf.set_text_color(0, 0, 0)
        title = test_data.get('title', 'Test Report')
        pdf.cell(0, 10, title, ln=True, align='C')

        pdf.set_font('Helvetica', '', 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align='C')
        pdf.ln(10)

        # Test Information
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_text_color(31, 119, 180)
        pdf.cell(0, 8, 'Test Information', ln=True)

        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(0, 0, 0)

        info_fields = [
            ('Protocol ID', test_data.get('protocol_id', 'N/A')),
            ('Protocol Name', test_data.get('protocol_name', 'N/A')),
            ('Sample ID', test_data.get('sample_id', 'N/A')),
            ('Test Date', test_data.get('test_date', datetime.now().strftime('%Y-%m-%d'))),
            ('Operator', test_data.get('operator', 'N/A')),
        ]

        for field, value in info_fields:
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(50, 7, f"{field}:", 0)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 7, str(value), ln=True)

        pdf.ln(10)

        # Results Summary
        if results_data:
            pdf.set_font('Helvetica', 'B', 12)
            pdf.set_text_color(31, 119, 180)
            pdf.cell(0, 8, 'Results Summary', ln=True)

            total = len(results_data)
            passed = sum(1 for r in results_data if r.get('status', '').upper() == 'PASS')
            failed = sum(1 for r in results_data if r.get('status', '').upper() == 'FAIL')

            pdf.set_font('Helvetica', '', 10)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(50, 7, f"Total Tests: {total}", ln=True)
            pdf.cell(50, 7, f"Passed: {passed}", ln=True)
            pdf.cell(50, 7, f"Failed: {failed}", ln=True)
            pdf.cell(50, 7, f"Pass Rate: {(passed/total*100):.1f}%" if total > 0 else "N/A", ln=True)

            pdf.ln(10)

            # Overall Status
            overall = 'FAIL' if failed > 0 else 'PASS'
            if overall == 'PASS':
                pdf.set_fill_color(40, 167, 69)
            else:
                pdf.set_fill_color(220, 53, 69)

            pdf.set_font('Helvetica', 'B', 14)
            pdf.set_text_color(255, 255, 255)
            pdf.cell(0, 12, f"OVERALL RESULT: {overall}", ln=True, align='C', fill=True)

        pdf.ln(20)

        # Signature Section
        pdf.set_font('Helvetica', '', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 8, 'Authorization:', ln=True)
        pdf.ln(5)
        pdf.cell(30, 8, 'Tested By:', 0)
        pdf.cell(60, 8, '_' * 25, 0)
        pdf.cell(20, 8, 'Date:', 0)
        pdf.cell(0, 8, '_' * 15, ln=True)

        pdf.cell(30, 8, 'Reviewed By:', 0)
        pdf.cell(60, 8, '_' * 25, 0)
        pdf.cell(20, 8, 'Date:', 0)
        pdf.cell(0, 8, '_' * 15, ln=True)

        pdf.cell(30, 8, 'Approved By:', 0)
        pdf.cell(60, 8, '_' * 25, 0)
        pdf.cell(20, 8, 'Date:', 0)
        pdf.cell(0, 8, '_' * 15, ln=True)

        # Return PDF bytes
        return bytes(pdf.output())


class ExcelReportGenerator:
    """Generate Excel reports with multiple sheets"""

    def __init__(self):
        self.company = CompanySettings.get_settings()

    def generate_test_report(
        self,
        test_data: Dict[str, Any],
        results_data: List[Dict[str, Any]],
        parameters_data: Optional[Dict[str, Any]] = None
    ) -> Optional[bytes]:
        """
        Generate Excel report with multiple sheets

        Sheets:
        1. Test Data - Raw test results
        2. Summary - Summary statistics
        3. Parameters - Test parameters and specifications
        """
        if not OPENPYXL_AVAILABLE:
            st.error("Excel generation requires openpyxl. Install with: pip install openpyxl")
            return None

        wb = Workbook()

        # Sheet 1: Test Data
        ws_data = wb.active
        ws_data.title = "Test Data"
        self._create_test_data_sheet(ws_data, test_data, results_data)

        # Sheet 2: Summary
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, test_data, results_data)

        # Sheet 3: Parameters
        ws_params = wb.create_sheet("Parameters")
        self._create_parameters_sheet(ws_params, test_data, parameters_data)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _create_test_data_sheet(
        self,
        ws,
        test_data: Dict[str, Any],
        results_data: List[Dict[str, Any]]
    ):
        """Create the test data sheet"""
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F77B4', end_color='1F77B4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws['A1'] = self.company['company_name']
        ws['A1'].font = Font(bold=True, size=14, color='FF6B35')
        ws.merge_cells('A1:G1')

        ws['A2'] = f"Test Report - {test_data.get('title', 'Test Results')}"
        ws['A2'].font = Font(bold=True, size=12)
        ws.merge_cells('A2:G2')

        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A3:G3')

        # Test Info
        row = 5
        info_fields = [
            ('Protocol ID', test_data.get('protocol_id', 'N/A')),
            ('Protocol Name', test_data.get('protocol_name', 'N/A')),
            ('Sample ID', test_data.get('sample_id', 'N/A')),
            ('Test Date', test_data.get('test_date', 'N/A')),
            ('Operator', test_data.get('operator', 'N/A')),
        ]

        for field, value in info_fields:
            ws.cell(row=row, column=1, value=field).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

        # Results Header
        headers = ['#', 'Parameter', 'Value', 'Unit', 'Spec Min', 'Spec Max', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Results Data
        pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        fail_fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')

        for i, result in enumerate(results_data, 1):
            row += 1
            ws.cell(row=row, column=1, value=i).border = thin_border
            ws.cell(row=row, column=2, value=result.get('parameter', '')).border = thin_border
            ws.cell(row=row, column=3, value=result.get('value', '')).border = thin_border
            ws.cell(row=row, column=4, value=result.get('unit', '')).border = thin_border
            ws.cell(row=row, column=5, value=result.get('spec_min', '')).border = thin_border
            ws.cell(row=row, column=6, value=result.get('spec_max', '')).border = thin_border

            status_cell = ws.cell(row=row, column=7, value=result.get('status', ''))
            status_cell.border = thin_border
            status_cell.alignment = Alignment(horizontal='center')

            status = result.get('status', '').upper()
            if status == 'PASS':
                status_cell.fill = pass_fill
            elif status == 'FAIL':
                status_cell.fill = fail_fill

        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 10

    def _create_summary_sheet(
        self,
        ws,
        test_data: Dict[str, Any],
        results_data: List[Dict[str, Any]]
    ):
        """Create the summary sheet"""
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='28A745', end_color='28A745', fill_type='solid')

        ws['A1'] = 'Test Summary Report'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        # Calculate statistics
        total = len(results_data) if results_data else 0
        passed = sum(1 for r in results_data if r.get('status', '').upper() == 'PASS') if results_data else 0
        failed = sum(1 for r in results_data if r.get('status', '').upper() == 'FAIL') if results_data else 0

        summary_data = [
            ['Metric', 'Value'],
            ['Protocol ID', test_data.get('protocol_id', 'N/A')],
            ['Protocol Name', test_data.get('protocol_name', 'N/A')],
            ['Sample ID', test_data.get('sample_id', 'N/A')],
            ['Test Date', test_data.get('test_date', 'N/A')],
            ['Total Measurements', total],
            ['Passed', passed],
            ['Failed', failed],
            ['Pass Rate', f"{(passed/total*100):.1f}%" if total > 0 else 'N/A'],
            ['Overall Status', 'FAIL' if failed > 0 else ('PASS' if total > 0 else 'N/A')]
        ]

        for row_idx, row_data in enumerate(summary_data, 3):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 3:
                    cell.font = header_font
                    cell.fill = header_fill
                elif col_idx == 1:
                    cell.font = Font(bold=True)

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30

    def _create_parameters_sheet(
        self,
        ws,
        test_data: Dict[str, Any],
        parameters_data: Optional[Dict[str, Any]]
    ):
        """Create the parameters sheet"""
        ws['A1'] = 'Test Parameters & Specifications'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:D1')

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')

        headers = ['Parameter', 'Specification', 'Unit', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Default parameters if none provided
        if not parameters_data:
            parameters_data = {
                'parameters': [
                    {'name': 'Temperature', 'spec': '25 +/- 2', 'unit': 'C', 'notes': 'Standard Test Conditions'},
                    {'name': 'Irradiance', 'spec': '1000', 'unit': 'W/m2', 'notes': 'STC'},
                    {'name': 'Humidity', 'spec': '< 75', 'unit': '%RH', 'notes': ''},
                ]
            }

        for row_idx, param in enumerate(parameters_data.get('parameters', []), 4):
            ws.cell(row=row_idx, column=1, value=param.get('name', ''))
            ws.cell(row=row_idx, column=2, value=param.get('spec', ''))
            ws.cell(row=row_idx, column=3, value=param.get('unit', ''))
            ws.cell(row=row_idx, column=4, value=param.get('notes', ''))

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 30


def get_report_generator(format_type: str = 'pdf'):
    """
    Factory function to get appropriate report generator

    Args:
        format_type: 'pdf', 'fpdf', or 'excel'

    Returns:
        Report generator instance
    """
    if format_type == 'pdf':
        return PDFReportGenerator()
    elif format_type == 'fpdf':
        return FPDFReportGenerator()
    elif format_type == 'excel':
        return ExcelReportGenerator()
    else:
        raise ValueError(f"Unknown format type: {format_type}")


def render_report_generator_ui():
    """Render the report generation UI in Streamlit"""
    st.subheader("Report Generator")

    # Check library availability
    col1, col2, col3 = st.columns(3)
    with col1:
        if REPORTLAB_AVAILABLE:
            st.success("ReportLab: Available")
        else:
            st.warning("ReportLab: Not installed")
    with col2:
        if FPDF_AVAILABLE:
            st.success("FPDF2: Available")
        else:
            st.warning("FPDF2: Not installed")
    with col3:
        if OPENPYXL_AVAILABLE:
            st.success("OpenPyXL: Available")
        else:
            st.warning("OpenPyXL: Not installed")

    st.divider()

    # Report Configuration
    col1, col2 = st.columns(2)

    with col1:
        report_title = st.text_input("Report Title", value="Solar PV Module Test Report")
        protocol_id = st.text_input("Protocol ID", value="P1-IV-001")
        protocol_name = st.selectbox("Protocol Name", [
            "P1 - I-V Performance Characterization",
            "P2 - Temperature Coefficient Test",
            "P3 - Low Irradiance Test",
            "P28 - Humidity Freeze Test",
            "P40 - Mechanical Load Test"
        ])

    with col2:
        sample_id = st.text_input("Sample ID", value="MOD-2024-001")
        test_date = st.date_input("Test Date")
        operator = st.text_input("Operator Name", value="Test Engineer")

    st.divider()

    # Sample Results Data
    st.subheader("Test Results")

    # Default sample data
    default_results = [
        {'parameter': 'Pmax', 'value': 405.2, 'unit': 'W', 'spec_min': 385, 'spec_max': 425, 'status': 'PASS'},
        {'parameter': 'Voc', 'value': 49.8, 'unit': 'V', 'spec_min': 48, 'spec_max': 52, 'status': 'PASS'},
        {'parameter': 'Isc', 'value': 10.2, 'unit': 'A', 'spec_min': 9.5, 'spec_max': 10.8, 'status': 'PASS'},
        {'parameter': 'Vmp', 'value': 41.5, 'unit': 'V', 'spec_min': 39, 'spec_max': 44, 'status': 'PASS'},
        {'parameter': 'Imp', 'value': 9.76, 'unit': 'A', 'spec_min': 9.2, 'spec_max': 10.2, 'status': 'PASS'},
        {'parameter': 'FF', 'value': 79.8, 'unit': '%', 'spec_min': 75, 'spec_max': 85, 'status': 'PASS'},
        {'parameter': 'Efficiency', 'value': 20.3, 'unit': '%', 'spec_min': 19, 'spec_max': 22, 'status': 'PASS'},
    ]

    results_df = pd.DataFrame(default_results)
    edited_df = st.data_editor(
        results_df,
        width="stretch",
        num_rows="dynamic"
    )

    st.divider()

    # Generate Reports
    st.subheader("Generate Report")

    col1, col2, col3 = st.columns(3)

    test_data = {
        'title': report_title,
        'protocol_id': protocol_id,
        'protocol_name': protocol_name,
        'sample_id': sample_id,
        'test_date': test_date.strftime('%Y-%m-%d') if test_date else datetime.now().strftime('%Y-%m-%d'),
        'operator': operator,
        'equipment': 'Solar Simulator SS-1000',
        'conditions': 'STC (25C, 1000 W/m2, AM1.5)'
    }

    results_data = edited_df.to_dict('records')

    with col1:
        if st.button("Generate PDF Report", type="primary", width="stretch"):
            if REPORTLAB_AVAILABLE:
                generator = PDFReportGenerator()
                pdf_bytes = generator.generate_test_report(test_data, results_data)
                if pdf_bytes:
                    st.download_button(
                        label="Download PDF",
                        data=pdf_bytes,
                        file_name=f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf"
                    )
                    st.success("PDF generated successfully!")
            elif FPDF_AVAILABLE:
                generator = FPDFReportGenerator()
                pdf_bytes = generator.generate_test_report(test_data, results_data)
                if pdf_bytes:
                    st.download_button(
                        label="Download PDF",
                        data=pdf_bytes,
                        file_name=f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf"
                    )
                    st.success("PDF generated successfully!")
            else:
                st.error("No PDF library available. Install reportlab or fpdf2.")

    with col2:
        if st.button("Generate Excel Report", width="stretch"):
            if OPENPYXL_AVAILABLE:
                generator = ExcelReportGenerator()
                excel_bytes = generator.generate_test_report(test_data, results_data)
                if excel_bytes:
                    st.download_button(
                        label="Download Excel",
                        data=excel_bytes,
                        file_name=f"test_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    st.success("Excel report generated successfully!")
            else:
                st.error("OpenPyXL not available. Install with: pip install openpyxl")

    with col3:
        if st.button("Preview Report", width="stretch"):
            st.info("Report Preview")

            # Summary statistics
            total = len(results_data)
            passed = sum(1 for r in results_data if r.get('status', '').upper() == 'PASS')
            failed = total - passed

            metrics_col1, metrics_col2, metrics_col3 = st.columns(3)
            with metrics_col1:
                st.metric("Total Tests", total)
            with metrics_col2:
                st.metric("Passed", passed)
            with metrics_col3:
                st.metric("Failed", failed)

            if failed > 0:
                st.error(f"OVERALL STATUS: FAIL")
            else:
                st.success(f"OVERALL STATUS: PASS")


# Export functions for use in other modules
__all__ = [
    'PDFReportGenerator',
    'FPDFReportGenerator',
    'ExcelReportGenerator',
    'CompanySettings',
    'get_report_generator',
    'render_report_generator_ui',
    'REPORTLAB_AVAILABLE',
    'FPDF_AVAILABLE',
    'OPENPYXL_AVAILABLE'
]
