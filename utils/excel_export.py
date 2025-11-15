"""
Excel Export Utilities
======================
Export audit data to Excel format using openpyxl.
"""

from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from config.settings import config
import os


def export_audit_to_excel(audit, output_path=None):
    """
    Export audit data to Excel

    Args:
        audit: Audit object
        output_path: Optional output file path

    Returns:
        Path to generated Excel file
    """
    if output_path is None:
        filename = f"audit_export_{audit.audit_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(config.REPORTS_DIR, filename)

    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Audit Information Sheet
    ws_info = wb.create_sheet("Audit Information")

    # Header style
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Audit Info
    ws_info['A1'] = 'Field'
    ws_info['B1'] = 'Value'
    ws_info['A1'].fill = header_fill
    ws_info['A1'].font = header_font
    ws_info['B1'].fill = header_fill
    ws_info['B1'].font = header_font

    info_data = [
        ['Audit Number', audit.audit_number],
        ['Audit Date', audit.actual_date.strftime('%Y-%m-%d') if audit.actual_date else 'N/A'],
        ['Auditee Entity', audit.schedule.auditee_entity.name if audit.schedule else 'N/A'],
        ['Audit Type', audit.schedule.audit_type.name if audit.schedule else 'N/A'],
        ['Lead Auditor', audit.schedule.auditor.full_name if audit.schedule else 'N/A'],
        ['Duration (hours)', audit.duration_hours or 0],
        ['Status', audit.status.value],
        ['Total Findings', audit.findings_count],
        ['NC Count', audit.nc_count],
        ['OFI Count', audit.ofi_count],
        ['Observations', audit.observations_count],
        ['Score', audit.score or 0],
    ]

    for idx, (field, value) in enumerate(info_data, start=2):
        ws_info[f'A{idx}'] = field
        ws_info[f'B{idx}'] = value
        ws_info[f'A{idx}'].font = Font(bold=True)

    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 40

    # NC/OFI Sheet
    if audit.nc_ofis:
        ws_nc = wb.create_sheet("NC/OFI")

        nc_data = []
        for nc in audit.nc_ofis:
            nc_data.append({
                'NC Number': nc.nc_number,
                'Type': nc.type.value.upper(),
                'Severity': nc.severity.value,
                'Category': nc.category or '',
                'Clause': nc.clause or '',
                'Description': nc.description,
                'Assignee': nc.assignee.full_name if nc.assignee else '',
                'Status': nc.status.value,
                'Due Date': nc.due_date.strftime('%Y-%m-%d') if nc.due_date else '',
                'Created': nc.created_at.strftime('%Y-%m-%d')
            })

        df = pd.DataFrame(nc_data)

        # Write dataframe to sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_nc.cell(row=r_idx, column=c_idx, value=value)

                # Header styling
                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjust column widths
        for column in ws_nc.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_nc.column_dimensions[column_letter].width = adjusted_width

    # Checklist Responses Sheet
    if audit.responses:
        ws_responses = wb.create_sheet("Checklist Responses")

        response_data = []
        for response in audit.responses:
            response_data.append({
                'Clause': response.checklist_item.clause_no,
                'Requirement': response.checklist_item.requirement,
                'Status': response.status.value,
                'Evidence': response.evidence or '',
                'Remarks': response.remarks or ''
            })

        df = pd.DataFrame(response_data)

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_responses.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in ws_responses.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 60)
            ws_responses.column_dimensions[column_letter].width = adjusted_width

    # Save workbook
    wb.save(output_path)

    return output_path


def export_nc_ofi_to_excel(nc_ofis, output_path=None):
    """
    Export NC/OFI list to Excel

    Args:
        nc_ofis: List of NC/OFI objects
        output_path: Optional output file path

    Returns:
        Path to generated Excel file
    """
    if output_path is None:
        filename = f"nc_ofi_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(config.REPORTS_DIR, filename)

    # Create dataframe
    data = []
    for nc in nc_ofis:
        data.append({
            'NC Number': nc.nc_number,
            'Type': nc.type.value.upper(),
            'Severity': nc.severity.value,
            'Category': nc.category or '',
            'Clause': nc.clause or '',
            'Description': nc.description,
            'Audit Number': nc.audit.audit_number if nc.audit else '',
            'Assignee': nc.assignee.full_name if nc.assignee else '',
            'Status': nc.status.value,
            'Due Date': nc.due_date.strftime('%Y-%m-%d') if nc.due_date else '',
            'Created': nc.created_at.strftime('%Y-%m-%d'),
            'Closed': nc.closed_at.strftime('%Y-%m-%d') if nc.closed_at else ''
        })

    df = pd.DataFrame(data)

    # Export to Excel
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='NC_OFI', index=False)

        # Get workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['NC_OFI']

        # Style header
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return output_path


def export_entities_to_excel(entities, output_path=None):
    """
    Export entities to Excel

    Args:
        entities: List of Entity objects
        output_path: Optional output file path

    Returns:
        Path to generated Excel file
    """
    if output_path is None:
        filename = f"entities_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = os.path.join(config.REPORTS_DIR, filename)

    data = []
    for entity in entities:
        data.append({
            'Code': entity.code,
            'Name': entity.name,
            'Type': entity.type,
            'Level': entity.level,
            'Location': entity.location or '',
            'Manager': entity.manager_name or '',
            'Email': entity.contact_email or '',
            'Phone': entity.contact_phone or '',
            'Parent': entity.parent.name if entity.parent else ''
        })

    df = pd.DataFrame(data)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Entities', index=False)

        workbook = writer.book
        worksheet = writer.sheets['Entities']

        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return output_path
